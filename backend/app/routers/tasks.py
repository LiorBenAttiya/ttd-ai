from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from typing import Optional
import io
import openpyxl
from app.core.security import get_current_user
from app.db.database import get_db
from app.models.task import TaskCreate, TaskUpdate, TaskReorderItem
from app.services import task_service
from sqlalchemy.ext.asyncio import AsyncSession

_COMPANY_PRIORITY = {"mep": 1, "lbatech": 1, "personal": 2}

router = APIRouter(prefix="/tasks", tags=["tasks"])


@router.get("")
async def list_tasks(
    view:       str = Query("kanban", pattern="^(kanban|list|timeline)$"),
    status:     Optional[str] = None,
    priority:   Optional[int] = None,
    category:   Optional[str] = Query(None, pattern="^(personal|business|all)$"),
    due_status: Optional[str] = None,
    mailbox_id: Optional[str] = None,
    search:     Optional[str] = None,
    limit:      int = Query(200, le=500),
    offset:     int = 0,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if view == "kanban":
        return await task_service.list_tasks_kanban(db, category=category)
    return await task_service.list_tasks(
        db, status=status, priority=priority, category=category,
        due_status=due_status, mailbox_id=mailbox_id,
        search=search, limit=limit, offset=offset,
    )


@router.post("", status_code=201)
async def create_task(
    payload: TaskCreate,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await task_service.create_task(
        db,
        description=payload.description,
        priority=int(payload.priority),
        due_date=payload.due_date,
        start_date=payload.start_date,
        source="web",
        mailbox_id=str(payload.mailbox_id) if payload.mailbox_id else None,
        contact_id=str(payload.contact_id) if payload.contact_id else None,
    )


@router.post("/import", status_code=200)
async def import_tasks_from_excel(
    file: UploadFile = File(...),
    confirm: bool = False,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    fname = (file.filename or "").lower()
    if not fname.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    raw = await file.read()
    try:
        # Try read_only first (faster), fall back to full load
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse Excel file: {e}")
    ws = wb["Tasks"] if "Tasks" in wb.sheetnames else wb.active
    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        task_val    = row[0] if len(row) > 0 else None
        type_val    = row[1] if len(row) > 1 else None
        company_val = row[2] if len(row) > 2 else None
        if not task_val or not str(task_val).strip():
            skipped += 1
            continue
        desc        = str(task_val).strip()
        task_type   = str(type_val).strip() if type_val else ""
        company_key = str(company_val).strip().lower() if company_val else ""
        priority    = _COMPANY_PRIORITY.get(company_key, 3)
        full_desc   = "[" + task_type + "] " + desc if task_type else desc
        rows.append({"description": full_desc, "priority": priority, "company_key": company_key})
    wb.close()
    if not confirm:
        summary = {"MEP": 0, "LBATECH": 0, "Personal": 0, "Other": 0}
        for r in rows:
            k = r["company_key"]
            if k == "mep":         summary["MEP"] += 1
            elif k == "lbatech":   summary["LBATECH"] += 1
            elif k == "personal":  summary["Personal"] += 1
            else:                  summary["Other"] += 1
        preview_rows = [{"description": r["description"], "priority": r["priority"]} for r in rows]
        return {"preview": True, "total": len(rows), "skipped": skipped, "summary": summary, "rows": preview_rows}
    inserted = 0
    errors   = 0
    for r in rows:
        try:
            await task_service.create_task(db, description=r["description"], priority=r["priority"], source="web")
            inserted += 1
        except Exception:
            errors += 1
    return {"preview": False, "inserted": inserted, "skipped": skipped, "errors": errors}


@router.patch("/reorder", status_code=200)
async def reorder_tasks(
    items: list[TaskReorderItem],
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await task_service.reorder_tasks(db, [i.model_dump() for i in items])
    return {"ok": True}


@router.get("/{task_id}")
async def get_task(
    task_id: str,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    task = await task_service.get_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


@router.patch("/{task_id}")
async def update_task(
    task_id: str,
    payload: TaskUpdate,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    updates = payload.model_dump(exclude_none=True)
    task = await task_service.update_task(db, task_id, updates)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


@router.delete("/{task_id}", status_code=204)
async def delete_task(
    task_id: str,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    deleted = await task_service.delete_task(db, task_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Task not found")


@router.post("/{task_id}/complete")
async def complete_task(
    task_id: str,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    task = await task_service.complete_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


@router.post("/{task_id}/archive")
async def archive_task(
    task_id: str,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    task = await task_service.update_task(db, task_id, {"status": "archived"})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


@router.post("/{task_id}/restore")
async def restore_task(
    task_id: str,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    task = await task_service.update_task(db, task_id, {"status": "pending"})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


@router.get("/{task_id}/messages")
async def get_task_messages(
    task_id: str,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await task_service.get_task_messages(db, task_id)


@router.get("/{task_id}/activity")
async def get_task_activity(
    task_id: str,
    _user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await task_service.get_task_activity(db, task_id)
